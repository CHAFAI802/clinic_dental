# stock/views.py
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, View
from django.urls import reverse_lazy
from django.http import HttpResponse
from openpyxl import Workbook
from .models import Product, Movement
from .forms import ProductForm, MovementForm
from .mixins import AdminRequiredMixin 
from django.http import HttpResponse
from django.views import View
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.comments import Comment
from datetime import datetime,timedelta,date
from .signals import apply_movement
from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render


# ---- Product Views ----

class ProductListView(AdminRequiredMixin, ListView):
    model = Product
    template_name = 'stock/product_list.html'
    context_object_name = 'products'
    paginate_by = 10

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        stock_colors = {}
        stock_variations = {}

        for product in context['products']:
            last_stock = product.last_stock or 0
            current_stock = product.current_stock or 0
            initial_stock = product.quantity
            if initial_stock == current_stock:
                variation = 100
            else:
                variation = current_stock  / initial_stock * 100

            stock_variations[product.id] = variation

            if variation > 50:
                stock_colors[product.id] = 'table-success'
            elif variation > 20:
                stock_colors[product.id] = 'table-warning'
            else:
                stock_colors[product.id] = 'table-danger'

        context['stock_colors'] = stock_colors
        context['stock_variations'] = stock_variations
        return context



class ProductDetailView(AdminRequiredMixin, DetailView):
    model = Product
    template_name = 'stock/product_detail.html'
    context_object_name = 'product'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        product = self.object

        # --- Variation ---
        initial_stock = product.quantity or 0
        current_stock = product.current_stock or 0
        last_stock = product.last_stock or 0

        if initial_stock == current_stock:
            variation = 100
        else:
            variation = current_stock  / initial_stock * 100

        context['stock_variation'] = variation

        # --- Couleur du tableau ---
        if variation < 20:
            context['stock_color'] = 'table-danger'
        elif variation < 50:
            context['stock_color'] = 'table-warning'
        else:
            context['stock_color'] = 'table-success'

        # --- Déterminer l’icône de flèche ---
        if current_stock > initial_stock:
            # flèche montante verte
            context['variation_icon'] = 'up'
        elif current_stock < initial_stock:
            # flèche descendante rouge
            context['variation_icon'] = 'down'
        else:
            context['variation_icon'] = 'equal'

        # --- Alertes expiration ---
        today = datetime.today().date()
        alert_messages = []

        if product.expiration_date and product.expiration_date < today:
            context['alert'] = True
            alert_messages.append(
                f"❌ Le produit '{product.name}' est déjà périmé depuis le "
                f"{product.expiration_date.strftime('%d/%m/%Y')} – il faut s’en débarrasser."
            )
        elif product.expiration_date and (product.expiration_date - today).days <= 15:
            context['alert'] = True
            alert_messages.append(
                f"⚠️ Attention : le produit '{product.name}' approche de sa date de péremption "
                f"({product.expiration_date.strftime('%d/%m/%Y')})."
            )

        if variation < 20:
            context['alert'] = True
            alert_messages.append(
                f"⚠️ Le stock est très bas ({variation:.1f}% du stock initial)."
            )
        elif variation < 50:
            context['alert'] = True
            alert_messages.append(
                f"⚠️ Le stock est inférieur à 50 % du stock initial  ({variation:.1f}%)."
            )

        if not alert_messages:
            context['alert'] = False
            context['alert_message'] = ''
        else:
            context['alert_message'] = "<br>".join(alert_messages)

        return context

class ProductCreateView(AdminRequiredMixin, CreateView):
    model = Product
    form_class = ProductForm
    template_name = 'stock/product_form.html'
    success_url = reverse_lazy('stock:product_list')

class ProductUpdateView(AdminRequiredMixin, UpdateView):
    model = Product
    form_class = ProductForm
    template_name = 'stock/product_form.html'
    success_url = reverse_lazy('stock:product_list')

class ProductDeleteView(AdminRequiredMixin, DeleteView):
    model = Product
    template_name = 'stock/product_confirm_delete.html'
    success_url = reverse_lazy('stock:product_list')

# ---- Movement Views ----

class MovementListView(AdminRequiredMixin, ListView):
    model = Movement
    template_name = 'stock/movement_list.html'
    context_object_name = 'movements'
    paginate_by = 15
    ordering = ['-date']

    def get_queryset(self):
        qs = super().get_queryset().select_related('product')
        product_id = self.request.GET.get('product')
        movement_type = self.request.GET.get('movement_type')

        if product_id:
            qs = qs.filter(product_id=product_id)
        if movement_type:
            qs = qs.filter(movement_type=movement_type)

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['all_products'] = Product.objects.all()
        # pour conserver les valeurs choisies dans le formulaire
        context['selected_product'] = self.request.GET.get('product', '')
        context['selected_movement_type'] = self.request.GET.get('movement_type', '')
        return context

class MovementCreateView(AdminRequiredMixin, CreateView):
    model = Movement
    form_class = MovementForm
    template_name = 'stock/movement_form.html'
    success_url = reverse_lazy('stock:product_list') 

    def form_valid(self, form):
        response = super().form_valid(form)
        movement = self.object
        apply_movement(movement.product, movement.movement_type, movement.movement_quantity)
        return response

class MovementUpdateView(AdminRequiredMixin, UpdateView):
    model = Movement
    form_class = MovementForm
    template_name = 'stock/movement_form.html'
    success_url = reverse_lazy('stock:movement_list')

class MovementDeleteView(AdminRequiredMixin, DeleteView):
    model = Movement
    template_name = 'stock/movement_confirm_delete.html'
    success_url = reverse_lazy('stock:movement_list')

class MovementDetailView(AdminRequiredMixin, DetailView):
    model = Movement
    template_name = 'stock/movement_detail.html'
    context_object_name = 'movement'

# ---- Export Excel ----


class MovementExportXLSXView(AdminRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        queryset = Movement.objects.select_related('product').order_by('date', 'id')

        wb = Workbook()
        ws = wb.active
        ws.title = "Mouvements de stock"

        headers = ['Date', 'Produit', 'Type', 'Quantité', 'Note', 'Stock restant', 'Expiration']
        ws.append(headers)

        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        # ajout bordure fine et alignement centré
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border  # bordure pour en-têtes

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        orange_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        stock_par_produit = {}

        for m in queryset:
            pid = m.product_id
            if pid not in stock_par_produit:
                stock_par_produit[pid] = m.product.current_stock

            if m.movement_type == 'IN':
                stock_par_produit[pid] -= m.movement_quantity
            else:
                stock_par_produit[pid] += m.movement_quantity

            stock_restant = stock_par_produit[pid]

            expiration = m.product.expiration_date
            today = datetime.today().date()
            alert = False
            comment_text = ""
            if expiration and (expiration - today).days <= 15:
                alert = True
                comment_text = f"⚠️ Produit proche de péremption : {expiration.strftime('%d/%m/%Y')}"

            row = [
                m.date.strftime("%d/%m/%Y %H:%M"),
                m.product.name,
                'Entrée' if m.movement_type == 'IN' else 'Sortie',
                m.movement_quantity,
                m.note or '',
                stock_restant,
                expiration.strftime("%d/%m/%Y") if expiration else ''
            ]
            ws.append(row)

            # Couleur Excel
            stock = m.product.current_stock
            if stock > 50:
                fill = green_fill
            elif stock < 20:
                fill = red_fill
            else:
                fill = orange_fill
            if alert:
                fill = orange_fill

            last_row_idx = ws.max_row
            for col_idx, cell in enumerate(ws[last_row_idx], 1):
                cell.fill = fill
                cell.border = thin_border           # ajout bordure
                cell.alignment = center_alignment   # ajout alignement centré
                if alert and col_idx == 2:
                    cell.comment = Comment(comment_text, "StockSystem")

        # Ajuster largeur colonnes
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="mouvements.xlsx"'
        wb.save(response)
        return response


def is_admin(user):
    return user.is_authenticated and user.role == "admin"


@login_required
@user_passes_test(is_admin)
def alerts_view(request):
    seuil_stock = 50
    seuil_jours = 15
    today = date.today()

    produits = Product.objects.all()

    produits_stock_faible = [p for p in produits if p.current_stock < seuil_stock]
    produits_peremption_proche = [
        p for p in produits
        if p.expiration_date and p.expiration_date <= today + timedelta(days=seuil_jours)
    ]

    context = {
        "produits_stock_faible": produits_stock_faible,
        "produits_peremption_proche": produits_peremption_proche,
    }
    return render(request, "stock/alerts_list.html", context)
